from selenium import webdriver
from time import sleep
import pprint
from selenium.webdriver.common.keys import Keys
import openpyxl
import datetime

#Program that compares your current followers with the followers you had last time you ran the program
# Notifies if any profile has disappeared in your follower listv(= Unfollowed you)
# Bug to be fixed: If a follower changes their username, they will appear as unfollowers


# related content for help on Page 257 of "Automating the boring stuff with Python"
browser = webdriver.Firefox(executable_path='/Users/a_saussier/Documents/geckodriver')

def login(username, password):

    # Open browser @ instagram
    browser.get('http://instagram.com')

    sleep(2)

    # find username field, send username key
    usernameElem = browser.find_element_by_name('username')
    usernameElem.send_keys(username)

    # Find password field, send password key
    passwordElem = browser.find_element_by_name('password')
    passwordElem.send_keys(password)

    # click on connexion button
    passwordElem.submit()

    # Time for the webpage to load
    sleep(10)

    # If the "Want to save your password?" page appears, click on 'Enregistrer vos identifiants plus tard'
    try:
        saveIDElem = browser.find_element_by_xpath('/html/body/div[1]/section/main/div/div/div/div/button')
        saveIDElem.click()
        sleep(5)
    except:
        sleep(5)

    # Click on 'Show notifications later'
    noNotifsElem = browser.find_element_by_xpath('/html/body/div[4]/div/div/div/div[3]/button[2]')
    noNotifsElem.click()
    sleep(5)


login('alexsaussier', 'Alexandre99$')

# Got to profile
toProfileElem = browser.find_element_by_xpath('/html/body/div[1]/section/main/section/div[3]/div[1]/div/div[2]/div[1]/a'
                                              )
toProfileElem.click()

sleep(4)

# Store number of followers in a variable from header on page
numFollowers = browser.find_element_by_xpath('/html/body/div[1]/section/main/div/header/section/ul/li[2]/a/span') \
    .get_attribute('title')
print('Number of followers: ' + numFollowers)

# click on followers to get scrolldown list menu
numFollowersElem = browser.find_element_by_xpath('/html/body/div[1]/section/main/div/header/section/ul/li[2]/a/span')
numFollowersElem.click()

sleep(1)

# Get a temporary list with the loaded followers
followerList = browser.find_elements_by_tag_name('a')
print(len(followerList))

sleep(5)
# Pass keys on first element of the list so that we can scroll down
scrollInFollowerBox = browser.find_element_by_tag_name('a')
scrollInFollowerBox.send_keys(Keys.DOWN)
sleep(2)

# Get all the followers

# While we don't have the total num of followers, scroll down and reset the variable with the newly loaded followers
# If the list doesn't get bigger after 5 scrolls, it means we reached the end
# 5 scrolls to leave some slack in case connexion is slow
followerListCounter = 0
while followerListCounter < 5:
    prev = len(followerList)
    scrollAgain = followerList[-1].send_keys(Keys.DOWN)
    sleep(1)
    followerList = browser.find_elements_by_tag_name('a')
    actual = len(followerList)
    print('loop ended, new followerlist length: ')
    print(len(followerList))
    if prev == actual:
        followerListCounter += 1

print('Process ended, all followers\' WebElements have been gathered')

followerNameList = []
for webElement in followerList:
    try:
        followerNameList.append(webElement.get_attribute('title'))
    except:
        print('Element of the class "%s" ignored' % (webElement.get_attribute('class')))

print('follower titles gathered')

# Remove blank elements from followerNameList
while '' in followerNameList:
    for username in followerNameList:
        if username == '':
            followerNameList.remove(username)


print('\n\n\nFinal List: ')
pprint.pprint(followerNameList)

# -------------------------- Write to excel ----------------------------------------

# Write list of followers to excel spreadsheet
wb = openpyxl.load_workbook('followerData.xlsx')
# Open sheet
sheet = wb['Followers']
# Set variables equal to last column and last row
highestCol = sheet.max_column
highestRow = sheet.max_row

# [A1] =  num of followers
sheet.cell(row=1, column=1).value = 'number of followers: '
# in [Ax] (new column) write the number of followers gathered by the program
sheet.cell(row=1, column=highestCol+1).value = numFollowers

# Same for date
sheet.cell(row=2, column=1).value = 'Date: '
sheet.cell(row=2, column=highestCol+1).value = datetime.datetime.now()

# Write followers in a column
for i in range(len(followerNameList)):
     sheet.cell(row=i+5, column=highestCol+1).value = str(followerNameList[i])

wb.save('followerData.xlsx')



# For debugging
print('Highest Column: ' + str(highestCol))
print('Highest row: ' + str(highestRow))

# Create two lists based of the two last columns
prevFollowers = []
newFollowers = []

for i in range(len(followerNameList)):
    # If a cell is empty, don't append to list
    if sheet.cell(row=5+i, column=highestCol).value == '' and sheet.cell(row=5+i, column=highestCol+1).value == '':
        continue
    if sheet.cell(row=5+i, column=highestCol).value == '':
        newFollowers.append(sheet.cell(row=5+i, column=highestCol+1).value)
    if sheet.cell(row=5+i, column=highestCol+1).value == '':
        prevFollowers.append(sheet.cell(row=5+i, column=highestCol).value)
    else:
        prevFollowers.append(sheet.cell(row=5+i, column=highestCol).value)
        newFollowers.append(sheet.cell(row=5+i, column=highestCol+1).value)

wb.save('followerData.xlsx')

unfollowerCount = 0
unfollowerList = []

for i in range(len(prevFollowers)):
    if prevFollowers[i] not in newFollowers:
        unfollowerList.append(prevFollowers[i])
        unfollowerCount += 1

if unfollowerCount == 0:
    print('Nobody unfollowed you since last run of the program')
else:
    print('Number of people who unfollowed: ' + str(unfollowerCount))
    pprint.pprint(unfollowerList)

# reset variables equal to last column and last row
highestCol = sheet.max_column
highestRow = sheet.max_row

sheet.cell(row=3, column=highestCol).value = 'Unfollowed by: ' + str(unfollowerCount)

for i in range(len(unfollowerList)):
    sheet.cell(row=highestRow+2+i, column=highestCol).value = unfollowerList[i]  # 'Highest' variables have been reset
wb.save('followerData.xlsx')

print('program end')
