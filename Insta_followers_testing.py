import openpyxl

wb = openpyxl.load_workbook('followerData.xlsx')
sheet = wb['Followers']
highestCol = sheet.max_column
highestRow = sheet.max_row
print('Highest Col: ' + str(highestCol))
print('Highest Row: ' + str(highestRow))

prevFollowers = []
newFollowers = []

for i in range(697):
    if sheet.cell(row=5 + i, column=highestCol - 1) is None:
        continue
    if sheet.cell(row=5 + i, column=highestCol) is None:
        continue
    else:
        prevFollowers.append(sheet.cell(row=5 + i, column=highestCol - 1).value)
        newFollowers.append(sheet.cell(row=5 + i, column=highestCol).value)

print('You have been unfollowed by:')

for i in range(len(prevFollowers)):
    if prevFollowers[i] not in newFollowers:
        print(prevFollowers[i])

print(highestCol)
print(highestRow)
wb.save('followerData.xlsx')